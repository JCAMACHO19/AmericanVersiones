import os
from openpyxl import load_workbook, Workbook
import re
import pandas as pd
from tkinter import Tk
from tkinter.filedialog import askopenfilename, asksaveasfilename
import download


# Función para cargar un archivo .xlsx
def cargar_archivo(mensaje):
    print(mensaje)
    Tk().withdraw()  # Cerrar la ventana raíz de Tkinter
    archivo_cargado = askopenfilename(title=mensaje, filetypes=[("Excel files", "*.xlsx")])
    if archivo_cargado.endswith('.xlsx'):
        return archivo_cargado
    raise ValueError("No se ha cargado un archivo con formato .xlsx")

# Solicitar los archivos al usuario
nombre_archivo_dian = cargar_archivo("Seleccionar archivo de DIAN")
nombre_archivo_sinco = cargar_archivo("Seleccionar archivo de SINCO")

# Procesar el archivo DIAN
df_dian = pd.read_excel(nombre_archivo_dian)
df_dian['Prefijo'] = df_dian['Prefijo'].fillna('')
df_dian['Folio'] = df_dian['Folio'].fillna('')
df_dian['Documento del Tercero'] = df_dian['Prefijo'].astype(str) + df_dian['Folio'].astype(str)
df_dian = df_dian.drop(columns=['Prefijo', 'Folio', 'NIT Receptor', 'Nombre Receptor'])
columnas_reordenadas_dian = ['Tipo de documento', 'CUFE/CUDE', 'Documento del Tercero', 'Fecha Emisión', 'Fecha Recepción', 'NIT Emisor', 'Nombre Emisor', 'IVA', 'ICA', 'IPC', 'Total', 'Estado', 'Grupo']
df_dian = df_dian[columnas_reordenadas_dian]
nombre_archivo_modificado_dian = "5.2. RELACION DIAN_modificado.xlsx"
df_dian.to_excel(nombre_archivo_modificado_dian, index=False)

# Procesar el archivo SINCO
wb = load_workbook(nombre_archivo_sinco)
hoja = wb.active
nuevo_wb = Workbook()
nueva_hoja = nuevo_wb.active
encabezados_encontrados = False
encabezados = []

for fila in hoja.iter_rows(values_only=True):
    if fila[0] == 'Tipo Registro':
        encabezados_encontrados = True
        encabezados = list(fila)
        tipo_doc_index = encabezados.index('Tipo Doc.')
        consecutivo_index = encabezados.index('Consecutivo')
        encabezados[tipo_doc_index] = 'Doc. Contable'
        del encabezados[consecutivo_index]
        nueva_hoja.append(encabezados)
        continue
    if encabezados_encontrados:
        cuenta_contable = fila[1]
        if not cuenta_contable.startswith(('539595')):
            nueva_fila = list(fila)
            doc_contable = f"{nueva_fila[tipo_doc_index]} {nueva_fila[consecutivo_index]}"
            nueva_fila[tipo_doc_index] = doc_contable
            del nueva_fila[consecutivo_index]
            documento_tercero_index = encabezados.index('Documento del Tercero')
            documento_tercero = nueva_fila[documento_tercero_index]
            if isinstance(documento_tercero, str):
                documento_tercero = re.sub(r'[-._ ]', '', documento_tercero)
            else:
                documento_tercero = str(documento_tercero) if documento_tercero is not None else ''
            nueva_fila[documento_tercero_index] = documento_tercero
            nueva_hoja.append(nueva_fila)

nuevo_nombre_archivo_sinco = "MovDocCuenta_filtrado.xlsx"
nuevo_wb.save(nuevo_nombre_archivo_sinco)

# Procesar el archivo generado de SINCO
df_sinco = pd.read_excel(nuevo_nombre_archivo_sinco)
df_sinco[['Numero de Factura', 'Concepto Restante']] = df_sinco['Concepto'].str.extract(r'([A-Z0-9\-]+)\s+([A-Z].*)')
df_sinco['Numero de Factura'] = df_sinco['Numero de Factura'].str.replace(r'[-_.]', '', regex=True)
df_sinco.drop(columns=['Concepto'], inplace=True)
nombre_archivo_tratado_sinco = "MovDocCuenta_tratado.xlsx"
df_sinco.to_excel(nombre_archivo_tratado_sinco, index=False)

# Fusionar y comparar los archivos
archivo1 = nombre_archivo_modificado_dian
archivo2 = nombre_archivo_tratado_sinco
archivo_salida = asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], title="Guardar archivo final como")

# Leer los archivos Excel
df1 = pd.read_excel(archivo1, sheet_name=None)  # Lee todas las hojas del archivo como un diccionario de DataFrames
df2 = pd.read_excel(archivo2, sheet_name=None)  # Lee todas las hojas del archivo como un diccionario de DataFrames

# Acceder a una hoja específica, por ejemplo 'Sheet1', si existe
df1_hoja1 = df1.get('Sheet1', None)
df2_hoja1 = df2.get('Sheet1', None)

if df1_hoja1 is not None and df2_hoja1 is not None:
    # Obtener las columnas relevantes
    documentos_tercero = df1_hoja1['Documento del Tercero']
    nit_emisor = df1_hoja1['NIT Emisor']
    numeros_factura = df2_hoja1['Numero de Factura']
    nit_receptor = df2_hoja1['NIT']
    fechas_contabilizacion = df2_hoja1['Fecha']
    docs_contables = df2_hoja1['Doc. Contable']
    documentos_tercero_2 = df2_hoja1['Documento del Tercero']

    # Crear una lista de valores para las nuevas columnas "Estado de Factura" y "Fecha Contabilización"
    estado_factura = []
    fechas_contabilizacion_nueva = []
    docs_contables_nuevos = []

    for doc, nit in zip(documentos_tercero, nit_emisor):
        # Verificar coincidencia en "Numero de Factura" y "Documento del Tercero"
        index_factura = (numeros_factura == doc) & (nit_receptor == nit)
        index_doc_tercero = (documentos_tercero_2 == doc) & (nit_receptor == nit)

        if any(index_factura) or any(index_doc_tercero):
            estado_factura.append('Contabilizado')
            if any(index_factura):
                fechas_contabilizacion_nueva.append(fechas_contabilizacion[index_factura].iloc[0] if index_factura.any() else None)
                docs_contables_nuevos.append(docs_contables[index_factura].iloc[0] if index_factura.any() else None)
            else:
                fechas_contabilizacion_nueva.append(fechas_contabilizacion[index_doc_tercero].iloc[0] if index_doc_tercero.any() else None)
                docs_contables_nuevos.append(docs_contables[index_doc_tercero].iloc[0] if index_doc_tercero.any() else None)
        else:
            estado_factura.append('Revisar')
            fechas_contabilizacion_nueva.append(None)
            docs_contables_nuevos.append(None)

    # Agregar las columnas "Estado de Factura", "Fecha Contabilización" y "Documento Contable" al DataFrame df1_hoja1
    df1_hoja1['Estado de Factura'] = estado_factura
    df1_hoja1['Fecha Contabilización'] = fechas_contabilizacion_nueva
    df1_hoja1['Documento Contable'] = docs_contables_nuevos

    # Guardar el DataFrame modificado en un nuevo archivo Excel
    with pd.ExcelWriter(archivo_salida, engine='openpyxl') as writer:
        df1_hoja1.to_excel(writer, sheet_name='Sheet1', index=False)

    print(f"Archivo final guardado como {archivo_salida}")

    # Eliminar archivos temporales
    archivos_temporales = [nombre_archivo_modificado_dian, nuevo_nombre_archivo_sinco, nombre_archivo_tratado_sinco]
    for archivo in archivos_temporales:
        try:
            os.remove(archivo)
        except Exception as e:
            pass  # Si hay algún error al eliminar el archivo, continuar sin mostrar mensaje

else:
    print("Una de las hojas no se pudo cargar correctamente.")